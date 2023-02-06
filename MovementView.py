__author__ = '720497'


import openpyxl
import time
from threading import Thread
import MovementConnect
import ClusterConnect
import Movement

threads = []
clustername = []

def run(sheet,stack_name,u,pwd):

    wb2= openpyxl.Workbook()
    sheet2 = wb2.active


    sheet2.title= sheet.title


    for j in range (1,12):
        sheet2.cell(1,j).value = sheet.cell(1,j).value

    f=1

    for i in range(2,sheet.max_row):
        if sheet.cell(i,1).value == stack_name:
           f=f+1
           for j in range (1,6):
            sheet2.cell(f,j).value = sheet.cell(i,j).value
           i=i+1
           while(sheet.cell(i,1).value == None ):
               f=f+1
               for j in range (1,6):
                sheet2.cell(f,j).value = sheet.cell(i,j).value
               i+=1
               if i==sheet.max_row+1:
                    break
           break



    clustername.append(sheet2.cell(2,2).value)
    clustername.append(sheet2.cell(2,3).value)


    result_srv = [{} for x in clustername]
    status_srv = [{} for x in clustername]

    result = {}
    status = {}




    for i in range(len(clustername)):
        time.sleep(2)
        process = Thread(target=MovementConnect.run, args=[clustername[i]+'.mmm.com',u,pwd,result_srv,status_srv,i])
        process.start()
        threads.append(process)

    for process in threads:
        process.join()

    ClusterConnect.run(clustername[0]+'.mmm.com',u,pwd,result,status,0)

    list = []



    for i in range (2,sheet2.max_row+1) :
        list.append(sheet2.cell(i,5).value +" "+ str(i))

    print(list)



    j=6

    for cluster_list in result_srv:

        for sid in list:
            sid_index = sid.split()
            arr_ascs = [sid_index[0],"SCS","sapstartsrv"]
            arr_ers = [sid_index[0],"ERS","sapstartsrv"]
            for entry in  cluster_list :



                if all(c in entry for c in arr_ascs):
                    sheet2.cell(int(sid_index[1]),j).value = "SCS:sapstartsrv present"  if (sheet2.cell(int(sid_index[1]),j).value == None) else sheet2.cell(int(sid_index[1]),j).value + "," + "SCS:sapstartsrv present"
                if all(c in entry for c in arr_ers):
                    sheet2.cell(int(sid_index[1]),j).value = "ERS:sapstartsrv present"  if (sheet2.cell(int(sid_index[1]),j).value == None) else sheet2.cell(int(sid_index[1]),j).value + "," + "ERS:sapstartsrv present"



        j=7





    sheet2.cell(2,4).value = status[0]
    sid_list = []
    for i in range (2,sheet2.max_row+1):
            sid_list.append(sheet2.cell(i,5).value +" "+ str(i))
    for cluster_service in result[0]:
        service=cluster_service.split()
        for sid in sid_list:
            sid_index=sid.split()
            if sid_index[0].lower() in service[0]:

                print(sid_index[0],service[0])
                index=int(sid_index[1])


                if int(service[1][-2:]) % 2 == 0:

                    sheet2.cell(index,9).value = service[0][8:]  if (sheet2.cell(index,9).value == None) else sheet2.cell(index,9).value + "," + service[0][8:]
                    sheet2.cell(index,11).value = service[2] if (sheet2.cell(index,11).value == None) else sheet2.cell(index,11).value + "," + service[2]


                else:

                    sheet2.cell(index,8).value = service[0][8:] if (sheet2.cell(index,8).value == None) else sheet2.cell(index,8).value + "," + service[0][8:]
                    sheet2.cell(index,10).value = service[2] if (sheet2.cell(index,10).value == None) else sheet2.cell(index,10).value + "," + service[2]







    t = time.localtime()
    timestamp = time.strftime('%b-%d-%Y_%H-%M', t)
    filename = ("Movement-"+sheet2.title+"-Status-" + timestamp + ".xlsx")
    wb2.save(filename)

    return(sheet2)

def run2(cluster_name,system_list,u,pwd):

    for i in range(len(system_list)):
        time.sleep(2)
        process = Thread(target=Movement.run1, args=[cluster_name+'.mmm.com',system_list[i],u,pwd])
        process.start()
        threads.append(process)

    for process in threads:
        process.join()


def run3(cluster_name,system_list,u,pwd):

    for i in range(len(system_list)):
        time.sleep(2)
        process = Thread(target=Movement.run2, args=[cluster_name+'.mmm.com',system_list[i],u,pwd])
        process.start()
        threads.append(process)

    for process in threads:
        process.join()











