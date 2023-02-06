from threading import Thread
import VIP_Connect
import openpyxl
import time

def run (sheet,u,pwd):
    threads = []
    link = []
    listab=[]

    wb2= openpyxl.Workbook()
    sheet2 = wb2.active
    row = sheet.max_row
    col= sheet.max_column

    sheet2.title= sheet.title

    for i in range(1,row+1):
        listab.append([])


    for r in range(1,row+1):
        for c in range(1,col+1):
            e=sheet.cell(r,c).value
            listab[r-1].append(e)



    for r in range(1,row+1):
        for c in range(1,col+1):
            j=sheet2.cell(row=r,column=c)
            j.value=listab[r-1][c-1]

    for i in range(2,sheet.max_row+1):
        link.append(sheet2.cell(i,2).value)
    print(link)

    return_code = [{} for x in link]


    for i in range(len(link)):
        time.sleep(2)
        process = Thread(target=VIP_Connect.run, args=[link[i],return_code,i])
        process.start()
        threads.append(process)

    for process in threads:
        process.join()

    h=0
    l=len(link)
    for i in range(2, sheet2.max_row+1):
        if(h<l):
         if sheet2.cell(i,2).value == link[h]:
            sheet2.cell(i,3).value = return_code[h]
            h=h+1

    t = time.localtime()
    timestamp = time.strftime('%b-%d-%Y_%H-%M', t)
    filename = ("VIP_Check_Result_"+ sheet2.title  +"-Status-" + timestamp + ".xlsx")
    wb2.save(filename)
    return(sheet2 )
