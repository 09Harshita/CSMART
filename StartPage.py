__author__ = '720497'

from SystemView import *
from tkinter import *
import openpyxl
from ToolTip import CreateToolTip
import MovementView
import ClusterView
import VIP_Check


logon = Tk()
v1 =StringVar()
logon.title("Login")

logon.configure(bg='white')
logon.geometry('200x175')
logon.resizable(0, 0)
#logon.resizable(True, True)

img = PhotoImage(file="C-SMART_LOGO.PNG")
label = Label(logon, image = img,bg='white')
label.grid(row=0,columnspan=3,sticky=NSEW)

l =Label(logon,text="Please login to continue",bg='white',font='Helvetica 10 bold')
l.grid(row=1,columnspan=3)

l1=Label(logon, text="Username",bg='white')
l1.grid(row=2)
e1=Entry(logon)
e1.grid(row=2, column=1)

l2=Label(logon, text="Password" ,bg='white')
l2.grid(row=3)
e2=Entry(logon,show='*')
e2.grid(row=3, column=1)

def f1(): #Getting Username and password
    u=e1.get()
    pwd=e2.get()
    if u=='' or pwd=='' or len(u) < 3 or len(u) > 15:
        popup = Toplevel(logon)
        popup.title("Error")
        popup.configure(bg='white')
        popup.geometry('200x100')
        l=Label(popup,text="Please enter your credentials",bg='white',font='Helvetica 10 bold')
        l.place(x=100, y=30, anchor="center")
        b=Button(popup,text="OK",command=popup.destroy)
        b.place(x=100,y=50, anchor="center")
    else:
        logon.withdraw()
        task(u,pwd)



b1=Button(logon, text="Log In",command=f1,font='Helvetica 10 bold')
b1.grid(row=6, columnspan=3)


def display(text,sheet,u,pwd): #Display server

    newwin2 = Toplevel(logon)
    newwin2.title("Server Overview")
    newwin2.configure(bg='white')
    newwin2.resizable(0, 0)

    label = Label(newwin2, image = img,bg='white')
    label.grid(row=0,columnspan=3,sticky=NSEW)

    b1 = Button(newwin2,text="Logout",width=15,font='Helvetica 10 bold',bg="navy",fg="white",command=logon.destroy)
    b1.grid(row=1,column=2,padx=2,pady=2)
    b2 = Button(newwin2,text="Stack Overview",width=15,font='Helvetica 10 bold',bg="navy",fg="white",command=lambda: [app(u,pwd,"System View"),newwin2.withdraw()])
    b2.grid(row=1,column=0,padx=2,pady=2)
    b3 = Button(newwin2,text="System Overview",width=15,font='Helvetica 10 bold',bg="navy",fg="white",command=lambda: [action(sheet,u,pwd),newwin2.withdraw()])
    b3.grid(row=1,column=1,padx=2,pady=2)
    label = Label(newwin2, text="Server Overview of " + text + " System",bg='white',font='Helvetica 10 bold')
    label.grid(row=2,columnspan=4)

    a=3
    b=0
    for i in range(1,sheet.max_row):
        if sheet.cell(i,1).value == text:
            if sheet.cell(i,4).value == "no" :
                x="red"
            elif sheet.cell(i,4).value == "hang":
                x="gray"
            else:
                x="green2"
            button=Button(newwin2,text=sheet.cell(i,2).value,bg=x,height=2,width=15,font='Helvetica 10 bold')
            button.grid(row=a,column=0,padx=2,pady=2)
            button1_ttp = CreateToolTip(button,sheet.cell(i,5).value )
            i+=1

            while(sheet.cell(i,1).value== None ):

                if sheet.cell(i,4).value == "no" :
                    x="red"
                elif sheet.cell(i,4).value == "hang":
                    x="gray"
                else:
                    x="green2"
                if b<2:
                     b=b+1
                else:
                    a=a+1
                    b=0
                button=Button(newwin2,text=sheet.cell(i,2).value,bg=x,height=2,width=15,font='Helvetica 10 bold')
                button.grid(row=a,column=b,padx=2,pady=2)
                button1_ttp = CreateToolTip(button,sheet.cell(i,5).value)
                i+=1
                if i==sheet.max_row+1:
                    break
            break

    label2 = Label(newwin2,text="Legend:",font='Helvetica 10 bold',bg="white",underline = True)
    label2.grid(row=a+1,column=0)
    button1 = Button(newwin2 ,text="Green",bg="green2",height=1,width=15,font='Helvetica 8 bold')
    button1.grid(row=a+2,column=0)

    l1 = Button(newwin2,text="Server available",bg="white",width=35,font='Helvetica 8 bold')
    l1.grid(row=a+2,column=1,columnspan=2,sticky=W )



    button3 = Button(newwin2 ,text="Red",bg="red",height=1,width=15,font='Helvetica 8 bold')
    button3.grid(row=a+3,column=0)
    l3 = Button(newwin2,text="Server not available",bg="white",width=35,font='Helvetica 8 bold')
    l3.grid(row=a+3,column=1,columnspan=2,sticky=W )

    button4 = Button(newwin2 ,text="Gray",bg="gray",height=1,width=15,font='Helvetica 8 bold')
    button4.grid(row=a+4,column=0)
    l4 = Button(newwin2,text="Technical issue or authenication fail",bg="white",width=35,font='Helvetica 8 bold')
    l4.grid(row=a+4,column=1,columnspan=2,sticky=W )

    l = Label(newwin2,text="",bg="white")
    l.grid(row=a+5,column=0)

    newwin2.protocol("WM_DELETE_WINDOW", logon.destroy)

def check(p,q,sheet): #Checking colour of System Button

    if sheet.cell(p+4,4).value == "no" or sheet.cell(p,4).value == "no" :
       x="red"
    elif sheet.cell(p+1,4).value == "no" or sheet.cell(p+2,4).value== "no" or sheet.cell(p+3,4).value  == "no" :
       x="yellow"
    else :
        for i in range(p+5,q):
            if sheet.cell(i,4).value == "no":
                x="yellow"
                break
            else:
                x="green2"

    return x

def hang(p,q,sheet):
    x=0
    for i in range (p,q):
        if sheet.cell(i,4).value == "hang":
            x=1
            break
    return x

def action(sheet,u,pwd): #Display System

    a=3
    b=0
    c=0
    arr = []
    newwin = Toplevel(logon)
    newwin.title("System Overview")
    newwin.configure(bg='white')
    newwin.resizable(0, 0)

    label = Label(newwin, image = img,bg='white')
    label.grid(row=0,columnspan=4,sticky=NSEW)

    b1 = Button(newwin,text="Logout",width=16,font='Helvetica 10 bold',bg="navy",fg="white",command=logon.destroy)
    b1.grid(row=1,column=2,columnspan=2,padx=2,pady=2)
    b2 = Button(newwin,text="Stack Overview",width=16,font='Helvetica 10 bold',bg="navy",fg="white",command=lambda: [app(u,pwd,"System View"),newwin.withdraw()])
    b2.grid(row=1,column=0,columnspan=2,padx=2,pady=2)

    label = Label(newwin, text="System Overview of " + sheet.title + " Stack",bg='white',font='Helvetica 10 bold')
    label.grid(row=2,columnspan=4)
    for i in range(2,sheet.max_row):
        if sheet.cell(i,1).value != None:
            arr.append(i)
    arr.append(sheet.max_row)



    for i in range(2,sheet.max_row+1):
        if sheet.cell(i, 1).value != None :
         p=arr[c]
         q=arr[c+1]
         h=hang(p,q,sheet)
         if h == 1:
            x="gray"
         else:
            x=check(p,q,sheet)
         button = Button(newwin ,text= sheet.cell(i, 1).value,bg=x,height=2,width=12,font='Helvetica 10 bold')
         button.config(command= lambda t=sheet.cell(i, 1).value, button = button: [display(t,sheet,u,pwd),newwin.withdraw()])
         button.grid(padx=2,pady=2,row=a,column=b)
         if b<3:
           b=b+1
         else:
            a=a+1
            b=0
         c=c+1

    label2 = Label(newwin,text="Legend:",font='Helvetica 10 bold',bg="white",underline = True)
    label2.grid(row=a+1,column=0)
    button1 = Button(newwin ,text="Green",bg="green2",height=1,width=12,font='Helvetica 8 bold')
    button1.grid(row=a+2,column=0)

    l1 = Button(newwin,text="All servers available",bg="white",width=16,font='Helvetica 8 bold')
    l1.grid(row=a+2,column=1)

    button2 = Button(newwin ,text="Yellow",bg="yellow",height=1,width=12,font='Helvetica 8 bold')
    button2.grid(row=a+3,column=0)
    l2 = Button(newwin,text="AS,ERS or Clus down",bg="white",width=16,font='Helvetica 8 bold')
    l2.grid(row=a+3,column=1)

    button3 = Button(newwin ,text="Red",bg="red",height=1,width=12,font='Helvetica 8 bold')
    button3.grid(row=a+2,column=2)
    l3 = Button(newwin,text="ASCS or Db down",bg="white",width=16,font='Helvetica 8 bold')
    l3.grid(row=a+2,column=3)

    button4 = Button(newwin ,text="Gray",bg="gray",height=1,width=12,font='Helvetica 8 bold')
    button4.grid(row=a+3,column=2)
    l4 = Button(newwin,text="Authenication fail",bg="white",width=16,font='Helvetica 8 bold')
    l4.grid(row=a+3,column=3)

    l = Label(newwin,text="",bg="white")
    l.grid(row=a+4,column=0)


    newwin.protocol("WM_DELETE_WINDOW", logon.destroy)


def addtolist(varList,c,sheet,u,pwd):

    List = []
    for item in varList:
        if item.get() != "":
            List.append(item.get())
    if len(List) == 0:
        print("please select system")
        popup = Toplevel(logon)
        popup.title("Error")
        popup.configure(bg='white')
        popup.geometry('200x100')
        l=Label(popup,text="Please select system",bg='white',font='Helvetica 10 bold')
        l.place(x=100, y=30, anchor="center")
        b=Button(popup,text="OK",command=popup.destroy)
        b.place(x=100,y=50, anchor="center")
    elif len(List) == c:

        flag = 0
        sheet2=reference(List,flag,sheet,u,pwd)
        action(sheet2,u,pwd)

        '''
        wb = openpyxl.load_workbook("System-DEV1-Status-May-02-2019_10-57.xlsx")
        sheet2 =wb.active
        action(sheet2,u,pwd)

        '''
    else:

        flag = 1
        sheet2=reference(List,flag,sheet,u,pwd)
        action(sheet2,u,pwd)
        '''
        wb = openpyxl.load_workbook("System-DEV1-Status-May-02-2019_10-57.xlsx")
        sheet2 =wb.active
        action(sheet2,u,pwd)
        '''
def select_all(cbuts):
    for i in cbuts:
        i.select()

def deselect_all(cbuts):
    for i in cbuts:
        i.deselect()

def status(sheet_name,u,pwd): #Taking status of server in excel sheet


    workbook = openpyxl.load_workbook("3M_Polaris_Server_List.xlsx")
    sheet = workbook[sheet_name]

    list = []
    cbuts = []
    varList = []
    for i in range (2,sheet.max_row) :
        if sheet.cell(i, 1).value != None :
            list.append(sheet.cell(i,1).value)
    org_count = len(list)
    a=3
    b=0
    newwin3 = Toplevel(logon)
    newwin3.title("System List")
    newwin3.configure(bg='white')
    newwin3.resizable(0, 0)

    b1 = Button(newwin3,text="Logout",width=15,font='Helvetica 10 bold',bg="navy",fg="white",command=logon.destroy)
    b1.grid(row=1,column=2,padx=2,pady=2,columnspan=2)
    b2 = Button(newwin3,text="Stack Overview",width=15,font='Helvetica 10 bold',bg="navy",fg="white",command=lambda: [app(u,pwd,"System View"),newwin3.withdraw()])
    b2.grid(row=1,column=0,padx=2,pady=2,columnspan=2)

    label = Label(newwin3, image = img,bg='white')
    label.grid(row=0,columnspan=4,sticky=NSEW)
    label2 = Label(newwin3, text="Select your system",bg='white',font='Helvetica 10 bold')
    label2.grid(row=2,columnspan=4)
    for index, item in enumerate(list):
        var1 = StringVar()
        cbuts.append(Checkbutton(newwin3, text=item,variable=var1,onvalue=item, offvalue="",bg='white',font='Helvetica 10 bold'))
        cbuts[index].grid(row=a, column=b, sticky=W)
        varList.append(var1)
        if b<3:
           b=b+1
        else:
            a=a+1
            b=0


    Button(newwin3, text = 'Select All',bg='white',font='Helvetica 10 bold', command =lambda: select_all(cbuts)).grid(row=a+1, column=0)
    Button(newwin3, text = 'Unselect All',bg='white',font='Helvetica 10 bold', command =lambda: deselect_all(cbuts)).grid(row=a+1, column=1)
    b1 = Button(newwin3, text="Check Status",font='Helvetica 10 bold', command=lambda: [addtolist(varList,org_count,sheet,u,pwd),newwin3.withdraw()])
    b1.grid(row=a+1, column=2,columnspan=4)

    l = Label(newwin3,text="",bg="white")
    l.grid(row=a+2,column=0)

    newwin3.protocol("WM_DELETE_WINDOW", logon.destroy)

def check_cluster(service_odd,service_even,s):
    if s != "Available":
        x= "gray"
    else :
        if(service_even == "No service running" and service_odd == "No service running"):
            x = "red"

        elif "scs" in service_odd and "ers" in service_even :
            x ="green2"
        else:
            x ="yellow"

    return x



def cluster(sheet_name,u,pwd):




    wb2 = openpyxl.load_workbook("ClusterList.xlsx")
    sheet = wb2[sheet_name]
    sheet2 = ClusterView.run(sheet,u,pwd)

    '''
    wb2 = openpyxl.load_workbook("Cluster-DEV1-Status-May-21-2019_15-43.xlsx")
    sheet2 = wb2.active
    '''


    newwin4 = Toplevel(logon)
    newwin4.title("Cluster View")
    newwin4.configure(background="white")
    newwin4.geometry("675x300")
    newwin4.resizable(0, 0)

    newwin4.grid_rowconfigure(0, weight=1)
    newwin4.grid_columnconfigure(0, weight=1)

    cnv = Canvas(newwin4,bg="white")
    cnv.grid(row=0, column=0, sticky='nswe')

    hScroll = Scrollbar(newwin4, orient=HORIZONTAL, command=cnv.xview)
    hScroll.grid(row=1, column=0, sticky='we')
    vScroll = Scrollbar(newwin4, orient=VERTICAL, command=cnv.yview)
    vScroll.grid(row=0, column=1, sticky='ns')
    cnv.configure(xscrollcommand=hScroll.set, yscrollcommand=vScroll.set)

    frm = Frame(cnv,bg="white")

    cnv.create_window(0, 0, window=frm, anchor='nw')





    label = Label(frm,bg='white', image = img)
    label.grid(row=0,column=0,columnspan=3,sticky=NSEW)

    b1 = Button(frm,text="Logout",width=16,font='Helvetica 10 bold',bg="navy",fg="white",command=logon.destroy)
    b1.grid(row=2,column=2,padx=2,pady=2)
    b2 = Button(frm,text="Stack Overview",width=16,font='Helvetica 10 bold',bg="navy",fg="white",command=lambda: [app(u,pwd,"Cluster View"),newwin4.withdraw()])
    b2.grid(row=2,column=1,padx=2,pady=2)

    label = Label(frm, text="Cluster Overview of " + sheet2.title + " Stack",bg='white',font='Helvetica 14 bold')
    label.grid(row=1,columnspan=3,sticky=NSEW)

    l1 = Button(frm,text = "SID",width=16,font='Helvetica 10 bold',fg = 'white',bg = 'black')
    l1.grid(row=3,column=0)

    l2 = Button(frm,text = "Odd Cluster",width=30,font='Helvetica 10 bold',fg = 'white',bg = 'black')
    l2.grid(row=3,column=1)

    l3 = Button(frm,text = "Even Cluster",width=30,font='Helvetica 10 bold',fg = 'white',bg = 'black')
    l3.grid(row=3,column=2)
    a=4
    for i in range (2,sheet2.max_row+1):

        if(sheet2.cell(i,1).value != None):
               cluster_odd = sheet2.cell(i,1).value
               cluster_even = sheet2.cell(i,2).value
               status = sheet2.cell(i,3).value

        service_odd = "No service running" if(sheet2.cell(i,5).value == None ) else sheet2.cell(i,5).value


        service_even = "No service running" if(sheet2.cell(i,6).value == None ) else sheet2.cell(i,6).value
        x=check_cluster(service_odd,service_even,status)

        b1 = Button(frm,text = sheet2.cell(i,4).value,width=16,font='Helvetica 10 bold',bg = x)
        b1.grid(row=a,column=0)

        b2 = Button(frm,text = cluster_odd.upper() +":"+service_odd,width=30,font='Helvetica 10 bold',bg = 'white')
        b2.grid(row=a,column=1)

        b3 = Button(frm,text = cluster_even.upper() +":"+service_even,font='Helvetica 10 bold',width=30,bg = 'white')
        b3.grid(row=a,column=2)
        a=a+1



    label2 = Label(frm,text="Legend:",font='Helvetica 10 bold',bg="white",underline = True)
    label2.grid(row=a+1,column=0)
    button1 = Button(frm ,text="Green",bg="green2",height=1,width=16,font='Helvetica 8 bold')
    button1.grid(row=a+2,column=0)

    l1 = Button(frm,text="ASCS running on odd cluster and ERS running on even cluster",bg="white",width=50,font='Helvetica 8 bold')
    l1.grid(row=a+2,column=1,columnspan=2, sticky='w')

    button2 = Button(frm ,text="Yellow",bg="yellow",height=1,width=16,font='Helvetica 8 bold')
    button2.grid(row=a+4,column=0)
    l2 = Button(frm,text="Either service not running on its respective cluster",bg="white",width=50,font='Helvetica 8 bold')
    l2.grid(row=a+4,column=1,columnspan=2, sticky='w')

    button3 = Button(frm ,text="Red",bg="red",height=1,width=16,font='Helvetica 8 bold')
    button3.grid(row=a+3,column=0)
    l3 = Button(frm,text="No service running on both the cluster",bg="white",width=50,font='Helvetica 8 bold')
    l3.grid(row=a+3,column=1,columnspan=2, sticky='w')

    button4 = Button(frm ,text="Gray",bg="gray",height=1,width=16,font='Helvetica 8 bold')
    button4.grid(row=a+5,column=0)
    l4 = Button(frm,text="Authentication failed",bg="white",width=50,font='Helvetica 8 bold')
    l4.grid(row=a+5,column=1,columnspan=2, sticky='w')

    l = Label(frm,text="",bg="white")
    l.grid(row=a+6,column=0)



    frm.update_idletasks()
    cnv.configure(scrollregion=(0, 0, frm.winfo_width(), frm.winfo_height()))

    newwin4.protocol("WM_DELETE_WINDOW", logon.destroy)


def movement(varList2,sheet2,u,pwd,type,stack_name,sheet_name):
    flag=""
    List = []
    arr =["ers","scs"]
    for item in varList2:
        if item.get() != "":
            List.append(item.get())

    if len(List) == 0:
        print("please select system")
        popup = Toplevel(logon)
        popup.title("Error")
        popup.configure(bg='white')
        popup.geometry('200x100')
        l=Label(popup,text="Please select system",bg='white',font='Helvetica 10 bold')
        l.place(x=100, y=30, anchor="center")
        b=Button(popup,text="OK",command=popup.destroy)
        b.place(x=100,y=50, anchor="center")
    print(List)

    target_odd = sheet2.cell(2,2).value
    target_even = sheet2.cell(2,3).value


    if type == "MTO":
        system1 =[]
        system = []
        system_list = []
        for sid in List:
            sid_index = sid.split()
            if "," in sheet2.cell(int(sid_index[1]),9).value :
                mylist = (sheet2.cell(int(sid_index[1]),9).value).split(',')
                for service in mylist :
                    system1.append(target_odd + " " + service)
                system_list.append(system1)
            else:
                service = sheet2.cell(int(sid_index[1]),9).value
                system.append(target_odd + " " + service)




    if type == "MTE":
        system1 =[]
        system = []
        system_list = []
        for sid in List:

            sid_index = sid.split()
            print(sid_index)
            if "," in sheet2.cell(int(sid_index[1]),8).value :
                mylist = (sheet2.cell(int(sid_index[1]),8).value).split(',')
                for service in mylist :
                    system1.append(target_even + " " + service)
                system_list.append(system1)

            else:
                service = sheet2.cell(int(sid_index[1]),8).value
                system.append(target_even + " " + service)






    elif type == "RB":
        system1 =[]
        system = []
        system_list = []
        for sid in List:
            sid_index = sid.split()
            print(sid_index)



            if all(c in str(sheet2.cell(int(sid_index[1]),9).value) for c in arr):
               mylist = (sheet2.cell(int(sid_index[1]),9).value).split(',')
               service = mylist[1] if "scs" in mylist[1] else mylist[0]
               system.append(target_odd +" "+ service)


            elif all(c in str(sheet2.cell(int(sid_index[1]),8).value) for c in arr):
                mylist = (sheet2.cell(int(sid_index[1]),8).value).split(',')
                service = mylist[1] if "ers" in mylist[1] else mylist[0]
                system.append(target_even +" "+ service)



            elif ((arr[0] in sheet2.cell(int(sid_index[1]),8).value) and (arr[1] in sheet2.cell(int(sid_index[1]),9).value)):
                service_scs = sheet2.cell(int(sid_index[1]),9).value
                service_ers = sheet2.cell(int(sid_index[1]),8).value

                system1.append(target_odd +" "+ service_scs)
                system1.append(target_even +" "+ service_ers)

                system_list.append(system1)


    if (system_list != [] and system != []) :
        print("Both")
        MovementView.run2(target_odd,system,u,pwd)
        MovementView.run3(target_odd,system_list,u,pwd)

    elif system != [] :
        print("run2")
        MovementView.run2(target_odd,system,u,pwd)

    elif system_list != []:
        print("run3")
        MovementView.run3(target_odd,system_list,u,pwd)




    cluster_check(stack_name,sheet_name,u,pwd)




def cluster_check(stack_name,sheet_name,u,pwd):


    wb2 = openpyxl.load_workbook("StackList.xlsx")
    sheet = wb2[sheet_name]
    print(stack_name)



    sheet2 = MovementView.run(sheet,stack_name,u,pwd)

    '''


    wb = openpyxl.load_workbook("Movement-SBX-Status-Aug-26-2019_11-45.xlsx")
    sheet2 = wb.active

    '''
    varList2 = []


    newwin4 = Toplevel(logon)
    newwin4.title("Cluster View")
    newwin4.configure(background="white")
    newwin4.geometry("1025x500")
    newwin4.resizable(0, 0)

    newwin4.grid_rowconfigure(0, weight=1)
    newwin4.grid_columnconfigure(0, weight=1)

    cnv = Canvas(newwin4,bg="white")
    cnv.grid(row=0, column=0, sticky='nswe')

    hScroll = Scrollbar(newwin4, orient=HORIZONTAL, command=cnv.xview)
    hScroll.grid(row=1, column=0, sticky='we')
    vScroll = Scrollbar(newwin4, orient=VERTICAL, command=cnv.yview)
    vScroll.grid(row=0, column=1, sticky='ns')
    cnv.configure(xscrollcommand=hScroll.set, yscrollcommand=vScroll.set)

    frm = Frame(cnv,bg="white")

    cnv.create_window(0, 0, window=frm, anchor='nw')

    label = Label(frm,bg='white', image = img)
    label.grid(row=0,column=0,columnspan=5,sticky=NSEW)

    b1 = Button(frm,text="Logout",width=16,font='Helvetica 10 bold',bg="navy",fg="white",command=logon.destroy)
    b1.grid(row=2,column=4,padx=2,pady=2)
    b2 = Button(frm,text="Stack Overview",width=16,font='Helvetica 10 bold',bg="navy",fg="white",command=lambda: [cluster_movement(sheet_name,u,pwd),newwin4.withdraw()])
    b2.grid(row=2,column=3,padx=2,pady=2)

    label = Label(frm, text="Cluster Overview of " + sheet2.title + " Stack",bg='white',font='Helvetica 14 bold')
    label.grid(row=1,columnspan=5,sticky=NSEW)


    l1 = Button(frm,text = "SID",width=16,font='Helvetica 10 bold',fg = 'white',bg = 'black')
    l1.grid(row=3,column=1)

    l4 = Button(frm,text = "SRV service status",width=30,font='Helvetica 10 bold',fg = 'white',bg = 'black')
    l4.grid(row=3,column=2)

    l2 = Button(frm,text = "Odd Cluster",width=30,font='Helvetica 10 bold',fg = 'white',bg = 'black')
    l2.grid(row=3,column=3)

    l3 = Button(frm,text = "Even Cluster",width=30,font='Helvetica 10 bold',fg = 'white',bg = 'black')
    l3.grid(row=3,column=4)

    a=4
    arr =["SCS","ERS"]
    checkbox = []
    for i in range (2,sheet2.max_row+1):
        var2 = StringVar()
        l=Checkbutton(frm,variable=var2,onvalue= sheet2.cell(i,5).value+" "+str(i), offvalue="",bg='white',font='Helvetica 10 bold')
        l.grid(row=a, column=0, sticky=NSEW)
        checkbox.append(l)
        varList2.append(var2)


        even_status = "Even:Both" if all(c in sheet2.cell(i,7).value for c in arr) else "Even:Missing"
        odd_status = "Odd:Both" if all(c in sheet2.cell(i,6).value for c in arr) else "Odd:Missing"
        srv_status = even_status +"&"+ odd_status



        cluster_odd = sheet2.cell(2,2).value
        cluster_even = sheet2.cell(2,3).value
        status = sheet2.cell(2,4).value

        service_odd = "No service running" if(sheet2.cell(i,8).value == None ) else sheet2.cell(i,8).value


        service_even = "No service running" if(sheet2.cell(i,9).value == None ) else sheet2.cell(i,9).value
        x=check_cluster(service_odd,service_even,status)

        b1 = Button(frm,text = sheet2.cell(i,5).value,width=16,font='Helvetica 10 bold',bg = x)
        b1.grid(row=a,column=1)

        b4 = Button(frm,text = srv_status ,width=30,font='Helvetica 10 bold',bg = 'white')
        b4.grid(row=a,column=2)

        b2 = Button(frm,text = cluster_odd.upper() +":"+service_odd,width=30,font='Helvetica 10 bold',bg = 'white')
        b2.grid(row=a,column=3)

        b3 = Button(frm,text = cluster_even.upper() +":"+service_even,font='Helvetica 10 bold',width=30,bg = 'white')
        b3.grid(row=a,column=4)



        a=a+1

    button4 = Button(frm,text = "Select All",font='Helvetica 10 bold',command =lambda: select_all(checkbox))
    button4.grid(row=3,column=0)



    button1 = Button(frm,text = "Move to ODD",width=16,font='Helvetica 10 bold',command=lambda: [movement(varList2,sheet2,u,pwd,"MTO",stack_name,sheet_name),newwin4.withdraw()])
    button1.grid(row=a+2,column=4)

    button2 = Button(frm,text = "Move to Even",width=16,font='Helvetica 10 bold',command=lambda: [movement(varList2,sheet2,u,pwd,"MTE",stack_name,sheet_name),newwin4.withdraw()])
    button2.grid(row=a+2,column=3)

    button3 = Button(frm,text = "Rebalance",width=16,font='Helvetica 10 bold',command=lambda: [movement(varList2,sheet2,u,pwd,"RB",stack_name,sheet_name),newwin4.withdraw()])
    button3.grid(row=a+2,column=2)


    label2 = Label(frm,text="Legend:",font='Helvetica 10 bold',bg="white",underline = True)
    label2.grid(row=a+3,column=1)
    button1 = Button(frm ,text="Green",bg="green2",height=1,width=16,font='Helvetica 8 bold')
    button1.grid(row=a+4,column=1)

    l1 = Button(frm,text="ASCS running on odd cluster and ERS running on even cluster",bg="white",width=50,font='Helvetica 8 bold')
    l1.grid(row=a+4,column=2,columnspan=3, sticky='w')

    button2 = Button(frm ,text="Yellow",bg="yellow",height=1,width=16,font='Helvetica 8 bold')
    button2.grid(row=a+6,column=1)
    l2 = Button(frm,text="Either service not running on its respective cluster",bg="white",width=50,font='Helvetica 8 bold')
    l2.grid(row=a+6,column=2,columnspan=3, sticky='w')

    button3 = Button(frm ,text="Red",bg="red",height=1,width=16,font='Helvetica 8 bold')
    button3.grid(row=a+5,column=1)
    l3 = Button(frm,text="No service running on both the cluster",bg="white",width=50,font='Helvetica 8 bold')
    l3.grid(row=a+5,column=2,columnspan=3, sticky='w')

    button4 = Button(frm ,text="Gray",bg="gray",height=1,width=16,font='Helvetica 8 bold')
    button4.grid(row=a+7,column=1)
    l4 = Button(frm,text="Authentication failed",bg="white",width=50,font='Helvetica 8 bold')
    l4.grid(row=a+7,column=2,columnspan=3, sticky='w')

    l = Label(frm,text="",bg="white")
    l.grid(row=a+8,column=0)



    frm.update_idletasks()
    cnv.configure(scrollregion=(0, 0, frm.winfo_width(), frm.winfo_height()))

    newwin4.protocol("WM_DELETE_WINDOW", logon.destroy)





def cluster_movement(sheet_name,u,pwd):

    wb2 = openpyxl.load_workbook("StackList.xlsx")
    sheet = wb2[sheet_name]

    a=3
    b=0

    newwin = Toplevel(logon)
    newwin.title("Stack Overview")
    newwin.configure(bg='white')
    newwin.resizable(0, 0)

    label = Label(newwin, image = img,bg='white')
    label.grid(row=0,columnspan=4,sticky=NSEW)

    b1 = Button(newwin,text="Logout",width=16,font='Helvetica 10 bold',bg="navy",fg="white",command=logon.destroy)
    b1.grid(row=1,column=2,columnspan=2,padx=2,pady=2)
    b2 = Button(newwin,text="Stack Overview",width=16,font='Helvetica 10 bold',bg="navy",fg="white",command=lambda: [app(u,pwd,"Cluster Movement"),newwin.withdraw()])
    b2.grid(row=1,column=0,columnspan=2,padx=2,pady=2)

    label = Label(newwin, text="Stack Overview of " + sheet.title + " Env",bg='white',font='Helvetica 10 bold')
    label.grid(row=2,columnspan=4)



    for i in range(2,sheet.max_row+1):
        if sheet.cell(i, 1).value != None :
         button = Button(newwin ,text= sheet.cell(i, 1).value,height=2,width=12,font='Helvetica 10 bold')
         button.config(command= lambda t=sheet.cell(i, 1).value, button = button: [cluster_check(t,sheet_name,u,pwd),newwin.withdraw()])
         button.grid(padx=2,pady=2,row=a,column=b)
         if b<3:
           b=b+1
         else:
            a=a+1
            b=0



    l = Label(newwin,text="",bg="white")
    l.grid(row=a+4,column=0)

def vip_check(sheet_name,u,pwd):
    wb2 = openpyxl.load_workbook("VIP_Check_Link_List.xlsx")
    sheet = wb2[sheet_name]
    sheet2 = VIP_Check.run(sheet, u, pwd)

    '''
    wb2 = openpyxl.load_workbook("Cluster-DEV1-Status-May-21-2019_15-43.xlsx")
    sheet2 = wb2.active
    '''

    newwin4 = Toplevel(logon)
    newwin4.title("VIP Check")
    newwin4.configure(background="white")
    newwin4.resizable(0, 0)

    newwin4.grid_rowconfigure(0, weight=1)
    newwin4.grid_columnconfigure(0, weight=1)

    cnv = Canvas(newwin4, bg="white")
    cnv.grid(row=0, column=0, sticky='nswe')

    hScroll = Scrollbar(newwin4, orient=HORIZONTAL, command=cnv.xview)
    hScroll.grid(row=1, column=0, sticky='we')
    vScroll = Scrollbar(newwin4, orient=VERTICAL, command=cnv.yview)
    vScroll.grid(row=0, column=1, sticky='ns')
    cnv.configure(xscrollcommand=hScroll.set, yscrollcommand=vScroll.set)

    frm = Frame(cnv, bg="white")

    cnv.create_window(0, 0, window=frm, anchor='nw')

    label = Label(frm, bg='white', image=img)
    label.grid(row=0, column=0, columnspan=2, sticky=NSEW)

    b1 = Button(frm, text="Logout", width=16, font='Helvetica 10 bold', bg="navy", fg="white", command=logon.destroy)
    b1.grid(row=2, column=1, padx=2, pady=2)
    b2 = Button(frm, text="Stack Overview", width=16, font='Helvetica 10 bold', bg="navy", fg="white",
                command=lambda: [app(u, pwd, "VIP Check"), newwin4.withdraw()])
    b2.grid(row=2, column=0, padx=2, pady=2)

    label = Label(frm, text="VIP Check of " + sheet2.title + " Stack", bg='white', font='Helvetica 14 bold')
    label.grid(row=1, column=0, columnspan=2, sticky=NSEW)

    l1 = Button(frm, text="VIP", width=40, font='Helvetica 10 bold', fg='white', bg='black')
    l1.grid(row=3, columnspan=2)
    a=4
    for i in range(2, sheet2.max_row + 1):
        if ((sheet2.cell(i, 3).value == 200) or (sheet2.cell(i, 3).value == 301) or (sheet2.cell(i, 3).value ==302)):
            x= "green2"
        elif(sheet2.cell(i, 3).value == 401):
            x= "yellow"
        else:
            x= "red"

        b1 = Button(frm, text=sheet2.cell(i, 1).value, width=40, font='Helvetica 10 bold', bg=x)
        b1.grid(row=a, columnspan=2)
        a=a+1


    label2 = Label(frm, text="Legend:", font='Helvetica 10 bold', bg="white", underline=True)
    label2.grid(row=a + 1, column=0)
    button1 = Button(frm, text="Green", bg="green2", height=1, width=20, font='Helvetica 8 bold')
    button1.grid(row=a + 2, column=0)

    l1 = Button(frm, text="VIP is accessible", bg="white", width=30,
                font='Helvetica 8 bold')
    l1.grid(row=a + 2, column=1, sticky='w')

    button1 = Button(frm, text="Yellow", bg="yellow", height=1, width=20, font='Helvetica 8 bold')
    button1.grid(row=a + 3, column=0)

    l1 = Button(frm, text="Requires Additional Authentication", bg="white", width=30,
                font='Helvetica 8 bold')
    l1.grid(row=a + 3, column=1, sticky='w')

    button2 = Button(frm, text="Red", bg="red", height=1, width=20, font='Helvetica 8 bold')
    button2.grid(row=a + 4, column=0)
    l2 = Button(frm, text="VIP not accessible", bg="white", width=30,
                font='Helvetica 8 bold')
    l2.grid(row=a + 4, column=1, sticky='w')


    l = Label(frm, text="", bg="white")
    l.grid(row=a + 5, column=0)

    frm.update_idletasks()
    cnv.configure(scrollregion=(0, 0, frm.winfo_width(), frm.winfo_height()))

    newwin4.protocol("WM_DELETE_WINDOW", logon.destroy)



def assignment(t,u,pwd,task_type):

    if task_type == "System View":
        status(t,u,pwd)

    elif task_type == "Cluster View":
        cluster(t,u,pwd)

    elif task_type == "Cluster Movement":
        cluster_movement(t,u,pwd)

    elif task_type == "VIP Check":
        vip_check(t,u,pwd)





def app(u,pwd,task_type):


    if task_type == "System View":
        wb_name = "3M_Polaris_Server_List.xlsx"


    elif task_type == "Cluster View":
        wb_name = "ClusterList.xlsx"
    elif task_type == "Cluster Movement":
        wb_name = "StackList.xlsx"
    elif task_type == "VIP Check":
        wb_name = "VIP_Check_Link_List.xlsx"

    workbook = openpyxl.load_workbook(wb_name)
    sheet_names = workbook.sheetnames
    res = len(workbook.sheetnames)

    print(res,sheet_names)

    root = Toplevel(logon)
    root.title("C-SMART")
    root.configure(bg='white')
    root.resizable(0, 0)

    a=4
    b=0
    label2 = Label(root, image = img,bg='white')
    label2.grid(row=0,columnspan=3,sticky=NSEW)

    b1 = Button(root,text="Change Task",width=12,bg="navy",fg="white",font='Helvetica 10 bold',command=lambda:[task(u,pwd),root.withdraw()])
    b1.grid(row=1,column=1,padx=5,pady=5)

    b2 = Button(root,text="Logout",width=12,bg="navy",fg="white",font='Helvetica 10 bold',command=logon.destroy)
    b2.grid(row=1,column=2,padx=5,pady=5)

    label1 = Label(root, text=task_type,bg='white',font='Helvetica 14 bold')
    label2 = Label(root, text="Select your environment:",bg='white',font='Helvetica 10 bold')
    label1.grid(row=2,columnspan=3)
    label2.grid(row=3,columnspan=3)

    for i in range (res) :

        button = Button(root,text=sheet_names[i],height=2, width=14,font='Helvetica 10 bold')
        button.config(command= lambda t=sheet_names[i], button = button: [assignment(t,u,pwd,task_type),root.withdraw()])
        button.grid(row=a,column=b)
        if b<2:
            b=b+1
        else:
            a=a+1
            b=0




    l = Label(root,text="",bg="white")
    l.grid(row=a+2,column=0)

    root.protocol("WM_DELETE_WINDOW", logon.destroy)





def task(u,pwd):

    task_win = Toplevel(logon)
    task_win.title("C-Smart")
    task_win.configure(bg='white')
    task_win.resizable(0, 0)

    label2 = Label(task_win, image = img,bg='white')
    label2.grid(row=0,columnspan=3,sticky=NSEW)


    b1 = Button(task_win,text="Logout",width=12,bg="navy",fg="white",font='Helvetica 10 bold',command=logon.destroy)
    b1.grid(row=1,column=2,padx=5,pady=5)


    label1 = Label(task_win, text="Welcome to C-SMART Application",bg='white',font='Helvetica 14 bold')
    label2 = Label(task_win, text="Select your task:",bg='white',font='Helvetica 10 bold')
    label1.grid(row=2,columnspan=3)
    label2.grid(row=3,columnspan=3)



    b1 = Button(task_win,text="System View",height=2,width=20,font='Helvetica 10 bold',command=lambda:[app(u,pwd,"System View"),task_win.withdraw()])
    b1.grid(row=4,padx=5,pady=5,columnspan=3,sticky=NSEW)

    b2 = Button(task_win,text="Cluster View",height=2,width=20,font='Helvetica 10 bold',command=lambda:[app(u,pwd,"Cluster View"),task_win.withdraw()])
    b2.grid(row=5,padx=5,pady=5,columnspan=3,sticky=NSEW)

    b2 = Button(task_win,text="Cluster Movement",height=2,width=20,font='Helvetica 10 bold',command=lambda:[app(u,pwd,"Cluster Movement"),task_win.withdraw()])
    b2.grid(row=6,padx=5,pady=5,columnspan=3,sticky=NSEW)

    b3 = Button(task_win, text="VIP Check", height=2, width=20, font='Helvetica 10 bold',command=lambda: [app(u, pwd, "VIP Check"), task_win.withdraw()])
    b3.grid(row=7, padx=5, pady=5, columnspan=3, sticky=NSEW)






logon.mainloop()