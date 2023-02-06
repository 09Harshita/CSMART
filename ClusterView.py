from threading import Thread
import ClusterConnect
import openpyxl
import time

def run (sheet,u,pwd):
    threads = []
    clustername = []
    listab=[]

    wb2= openpyxl.Workbook()
    sheet2 = wb2.active
    row = sheet.max_row
    col= sheet.max_column

    sheet2.title= sheet.title

    for i in range(1,row+1):
        listab.append([])


    for r in range(1,row+1):
        for c in range(1,col+1):
            e=sheet.cell(r,c).value
            listab[r-1].append(e)



    for r in range(1,row+1):
        for c in range(1,col+1):
            j=sheet2.cell(row=r,column=c)
            j.value=listab[r-1][c-1]

    for i in range(2,sheet.max_row+1):
        if sheet2.cell(i,1).value != None:
            clustername.append(sheet2.cell(i,1).value)
    print(clustername)

    result = [{} for x in clustername]
    status = [{} for x in clustername]

    for i in range(len(clustername)):
        time.sleep(2)
        process = Thread(target=ClusterConnect.run, args=[clustername[i]+'.mmm.com',u,pwd,result,status,i])
        process.start()
        threads.append(process)

    for process in threads:
        process.join()

    h=0
    l=len(clustername)
    for i in range(2, sheet2.max_row+1):
        if(h<l):
         if sheet2.cell(i,1).value == clustername[h]:
            sheet2.cell(i,3).value = status[h]
            h=h+1





    h=0
    for cluster_list in result:
        sid_list = []
        for i in range (2,sheet2.max_row+1):
            if sheet2.cell(i,1).value == clustername[h] :
                sid_list.append(sheet2.cell(i,4).value +" "+ str(i))
                i=i+1
                if (i==sheet2.max_row+1):
                        break
                while(sheet2.cell(i,1).value == None):
                    sid_list.append(sheet2.cell(i,4).value +" "+ str(i))
                    i=i+1
                    if (i==sheet2.max_row+1):
                        break
                break
        for cluster_service in cluster_list:
            service=cluster_service.split()
            for sid in sid_list:
                sid_index=sid.split()
                if sid_index[0].lower() in service[0]:

                    print(sid_index[0],service[0])
                    index=int(sid_index[1])
                    #print(service[1][-2:])

                    if int(service[1][-2:]) % 2 == 0:

                        sheet2.cell(index,6).value = service[0][8:]  if (sheet2.cell(index,6).value == None) else sheet2.cell(index,6).value + "," + service[0][8:]
                        sheet2.cell(index,8).value = service[2] if (sheet2.cell(index,8).value == None) else sheet2.cell(index,8).value + "," + service[2]


                    else:

                        sheet2.cell(index,5).value = service[0][8:] if (sheet2.cell(index,5).value == None) else sheet2.cell(index,5).value + "," + service[0][8:]
                        sheet2.cell(index,7).value = service[2] if (sheet2.cell(index,7).value == None) else sheet2.cell(index,7).value + "," + service[2]

        h=h+1




    t = time.localtime()
    timestamp = time.strftime('%b-%d-%Y_%H-%M', t)
    filename = ("Cluster-"+ sheet2.title  +"-Status-" + timestamp + ".xlsx")
    wb2.save(filename)
    return(sheet2 )
