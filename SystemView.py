__author__ = '720497'
import datetime
import ServerView
import openpyxl
import time
from threading import Thread

servername = []
threads = []


def reference(list,flag,sheet,u,p): # Connecting each server of SID to SSH
    wb2= openpyxl.Workbook()
    sheet2 = wb2.active

    listab=[]
    sheet2.title= sheet.title
    if flag == 0:
        print (list)
        row = sheet.max_row
        col= sheet.max_column



        for i in range(1,row+1):
            listab.append([])


        for r in range(1,row+1):
            for c in range(1,col+1):
                e=sheet.cell(r,c).value
                listab[r-1].append(e)



        for r in range(1,row+1):
            for c in range(1,col+1):
                j=sheet2.cell(row=r,column=c)
                j.value=listab[r-1][c-1]
    else :
        print (list)
        for j in range (1,6):
            sheet2.cell(1,j).value = sheet.cell(1,j).value

        f=1
        for l in list:
            for i in range(2,sheet.max_row):
                if sheet.cell(i,1).value == l:
                   f=f+1
                   for j in range (1,3):
                    sheet2.cell(f,j).value = sheet.cell(i,j).value
                   i=i+1
                   while(sheet.cell(i,1).value== None ):
                       f=f+1
                       for j in range (1,3):
                        sheet2.cell(f,2).value = sheet.cell(i,j).value
                       i+=1
                       if i==sheet.max_row+1:
                            break
                   break
    t1=datetime.datetime.now()

    for i in range(len(list)):
            time.sleep(2)
            process = Thread(target=ServerView.run, args=[sheet2,list[i],u,p])
            process.start()
            threads.append(process)

    for process in threads:
        process.join()


    t2= datetime.datetime.now()
    print("SystemView"+str(t2-t1))


    t = time.localtime()
    timestamp = time.strftime('%b-%d-%Y_%H-%M', t)
    filename = ("System-"+sheet2.title+"-Status-" + timestamp + ".xlsx")
    wb2.save(filename)
    return(sheet2)


