__author__ = '720497'

import paramiko
import time
import openpyxl



def run(servername,user,pwd,x,result,OsResult,index,run_service,lock): #SSH Connection

    wb =openpyxl.load_workbook('Services.xlsx')
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    resp = str()

    try:

        client.connect(servername, username=user, password=pwd)

        OsResult[index]="Available"
        lock.acquire()
        print(servername+":Host Available")
        lock.release()
        channel = client.invoke_shell()

        sn =servername.split('.')

        if x[1] == 'C' or x[1] == 'S':
            if x[2] == '1' or x[2]== '2' or x[2]== '3' or x[2]== '5' :
                name_suffix = 'ABAP_'
            if x[2] == '0' or x[2]=='6':
                name_suffix = 'JAVA_'
        elif x[1] == 'D' or x[1] == 'I' or x[1] == 'J' or x[1] == 'K' or x[1] == 'N' or x[1] == 'O' or x[1] == 'P' or x[1] == 'Q' or x[1] == 'Y' or x[1] == '5' or x[1] == '4' :
            name_suffix = 'JAVA_'
        elif x[1] == 'A' or x[1] == 'B' or x[1] == 'E' or x[1] == 'G' or x[1] == 'H' or x[1] == 'L' or x[1] == 'M' or x[1] == 'R' or x[1] == 'T' or x[1] == 'U' or x[1] == 'V' or x[1] == 'W' or x[1] == 'X' or x[1] == '3' or x[1] == '6' or x[1] == '8' :
            name_suffix = 'ABAP_'


        if sn[0][-3:] == "scs" :
            sheet_name=name_suffix +"ASCS"
        elif sn[0][-3:] == "ers" :
            sheet_name=name_suffix +"ERS"
        elif servername[5] == 'c' and servername[6] == 'l':
                sheet_name=name_suffix +"Cluster"
        elif servername[6] == 'c' and servername[7] == 'l':
                sheet_name=name_suffix +"Cluster"
        elif sn[0][-3:] == "hdb":
            sheet_name="HANA"
        elif sn[0][-2:] == "db":
            sheet_name="DB2"


        else:
            sheet_name = name_suffix +"AS"



        sheet=wb[sheet_name]




        for i in range (2,sheet.max_row+1):
            if sheet.cell(i,1).value !=None :
                service = []
                user = (sheet.cell(i,3).value).replace("<SID>",x).lower()
                command = (sheet.cell(i,1).value).replace("<user>",user)
                service_name = (sheet.cell(i,2).value).replace("<SID>",x)
                service.append(service_name)
                i+=1
                while(i <= sheet.max_row and sheet.cell(i,1).value == None ):
                    service_name = (sheet.cell(i,2).value).replace("<SID>",x)
                    service.append(service_name)
                    i+=1
                    if(i == sheet.max_row+1):
                        break




        command2 =(command + " \n")

        time.sleep(1)
        channel.send(command2)
        time.sleep(5)







        text=channel.recv(99999).decode("utf-8")
        r=""
        nr=""
        flag="no"


        if servername[5] == 'c' and servername[6] == 'l':
            for s in service:
                if s in text:
                   r = r + s + ","
                   flag="yes"
                   time.sleep(2)
                else:
                   nr = nr + s + ","
                   time.sleep(2)


        elif servername[6] == 'c' and servername[7] == 'l':
            for s in service:
                if s in text:
                   r = r + s + ","
                   flag="yes"
                   time.sleep(2)
                else:
                   nr = nr + s + ","
                   time.sleep(2)
        else:
            for s in service:
                if s in text:
                   r = r + s + ","
                   flag="yes"
                   time.sleep(2)
                else:
                   flag="no"
                   nr = nr +s + ","
                   time.sleep(2)

        time.sleep(10)

        if "ms.sap" in r :
            if "en.sap" in r:
                flag ="yes"
            else:
                flag ="no"
        if "en.sap" in r :
            if "ms.sap" in r:
                flag ="yes"
            else:
                flag ="no"

        if flag == "no":
            l =  "Service Not Running:" + nr[:-1]
        else:
            l =  "Service Running:" + r[:-1]


        run_service[index]=l
        result[index]=flag
        client.close()
        time.sleep(5)





    except paramiko.AuthenticationException:
        print(servername+":Authentication failed, please verify your credentials")
        result[index]= 'hang'
        run_service[index]="Authentication failed"
        OsResult[index]="Unavailable"
        client.close()
        time.sleep(1)
    except paramiko.BadHostKeyException as badHostKeyException:
        print(servername+":Unable to verify server's host key")
        result[index]= 'no'
        OsResult[index]="Unavailable"
        run_service[index]="Technical Error"
        client.close()
        time.sleep(1)
    except paramiko.SSHException:
        print(servername+":Unable to establish SSH connection")
        result[index]= 'no'
        OsResult[index]="Unavailable"
        run_service[index]="Unable to establish SSH connection"
        client.close()
        time.sleep(1)
    except ValueError as e:
        print(servername,e.args)
        result[index]= 'hang'
        OsResult[index]="Unavailable"
        run_service[index]="Technical Error"
        client.close()
        time.sleep(1)
    except Exception as e:
        print(servername,e.args)
        result[index]= 'no'
        OsResult[index]="Unavailable"
        run_service[index]="Technical Error"
        client.close()
        time.sleep(1)